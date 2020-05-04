import numpy as np
import pandas as pd
import os
import openpyxl
import shutil
from openpyxl.chart import BarChart, Reference, Series
from physicoModule import config
from physicoModule.physico_style import boldFont, centetAlignment, normal_box, name_box, topFill, frontFill, bottomFill, reserveFill
from physicoModule.physico_control import PhysicoControl, PhysicoMatch

# base_path = config.ROOT_CONFIG['base']
input_path = config.FOLDER_CONFIG['input']
output_path = config.FOLDER_CONFIG['output']
backup_path = config.FOLDER_CONFIG['backup']
result_day = config.FILE_CONFIG['day']
result_player = config.FILE_CONFIG['player']
result_match = config.FILE_CONFIG['match']


class ExcelWrite(PhysicoControl):
    DAYPOSITION_COLUMNS = ['Position', 'TR Time', 'Total Dist.', 'Dist. per min', 'HSR+Sprint', 'MSR', 'HSR', 'Sprint',
                           'Accel Cnt.', 'Decel Cnt.', 'Max Speed', 'GPS PL', 'Load']
    DAYPLAYER_COLUMNS = ['No.', 'Name', 'Position', 'Type', 'RPE', 'TR Time', 'Total Dist.', 'Dist. per min', 'HSR+Sprint', 'MSR', 'HSR', 'Sprint',
                         'Accel Cnt.', 'Decel Cnt.', 'Max Speed', 'GPS PL', 'Load', 'Mono', 'Strain', 'EWAM']
    PLAYER_COLUMNS = ['Camp', 'Date', 'Type', 'Type Info.', 'RPE', 'TR Time', 'Total Dist.', 'Dist. per min', 'HSR+Sprint', 'MSR', 'MSR %', 'HSR', 'HSR %', 'Sprint', 'Sprint %',
                      'Accel Cnt.', 'Decel Cnt.', 'Max Speed', 'GPS PL', 'Load', 'Mono', 'Strain', 'EWAM', 'Injury', 'Memo', 'Age', 'Sleep Time', 'Sleep', 'Tired', 'Stressed', 'Muscle', 'Height', 'Weight', 'Body Muscle', 'Body Fat']

    def __init__(self, PMANAGE):
        super().__init__(PMANAGE)
        self.output_excel_path = os.path.join(self.base_path, output_path)
        self.styler = ExcelStyle()
        self.chart = ExcelChart()

    def writePlayerExcel(self, name='ALL'):
        player_excel = self.makePlayerExcelData()
        team_data = player_excel['Team']
        del player_excel['Team']

        write_row = 2
        if name == 'ALL':
            file_path = os.path.join(self.output_excel_path, result_player)
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            team_data.to_excel(writer, sheet_name='Team', columns=team_data.columns,
                               index=False, float_format='%.1f', startrow=write_row)
            for key, value in player_excel.items():
                value = value[self.PLAYER_COLUMNS]
                value.to_excel(writer, sheet_name=key, columns=value.columns,
                               index=False, float_format='%.1f', startrow=write_row)
        else:
            file_path = os.path.join(
                self.output_excel_path, config.RESULT_CONFIG['player'], 'result_of_{}.xlsx'.format(name))
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            value = player_excel[name][self.PLAYER_COLUMNS]
            value.to_excel(writer, sheet_name=name, columns=value.columns,
                           index=False, float_format='%.1f', startrow=write_row)
        writer.save()
        writer.close()
        self.styler.stylePlayerExcel(file_path)
        return None

    def writeDayExcel(self, date='ALL'):
        file_set = self.file_set
        position_row = 13
        player_row = 23

        if date == 'ALL':
            file_path = os.path.join(
                self.output_excel_path, result_day)
            if os.path.isfile(file_path):
                os.remove(file_path)

            for file_date in file_set.keys():
                day_excel = self.makeDayExcelData(int(file_date))
                wtype, position, day = day_excel['type'], day_excel['position'][
                    self.DAYPOSITION_COLUMNS], day_excel['day'][self.DAYPLAYER_COLUMNS]
                day_type = file_date

                if os.path.isfile(file_path):
                    book = openpyxl.load_workbook(file_path)
                    writer = pd.ExcelWriter(
                        file_path, engine='openpyxl')
                    writer.book = book
                    position.to_excel(writer, sheet_name=day_type, index=False,
                                      float_format='%.1f', startrow=position_row, startcol=4)
                    day.to_excel(writer, sheet_name=day_type,
                                 index=False, float_format='%.1f', startrow=player_row)
                    writer.save()
                    writer.close()
                else:
                    writer = pd.ExcelWriter(
                        file_path, engine='openpyxl')
                    position.to_excel(writer, sheet_name=day_type, index=False,
                                      float_format='%.1f', startrow=position_row, startcol=4)
                    day.to_excel(writer, sheet_name=day_type,
                                 index=False, float_format='%.1f', startrow=player_row)
                    writer.save()
                    writer.close()
        else:
            day_excel = self.makeDayExcelData(date)
            file_path = os.path.join(
                self.output_excel_path, config.RESULT_CONFIG['day'], 'result_of_{}.xlsx'.format(date))
            wtype, position, day = day_excel['type'], day_excel['position'][
                self.DAYPOSITION_COLUMNS], day_excel['day'][self.DAYPLAYER_COLUMNS]
            day_type = str(date)+'_'+wtype
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            position.to_excel(writer, sheet_name=day_type, index=False,
                              float_format='%.1f', startrow=position_row, startcol=4)
            day.to_excel(writer, sheet_name=day_type,
                         index=False, float_format='%.1f', startrow=player_row)
            writer.save()
            writer.close()
        self.styler.styleDayExcel(file_path)
        self.chart.positionBar(file_path)
        return None

    def copyResultFile(self, copy_path):
        shutil.copy(os.path.join(
            self.base_path, output_path, result_player), copy_path)
        shutil.copy(os.path.join(
            self.base_path, output_path, result_day), copy_path)


class ExcelMatch(PhysicoMatch):
    MATCHPOSITION_COLUMNS = ['Position', 'TR Time', 'Total Dist.', 'Dist. per min', 'HSR+Sprint', 'MSR', 'HSR', 'Sprint',
                             'Accel Cnt.', 'Decel Cnt.', 'Max Speed', 'GPS PL', 'Load']
    MATCHPLAYER_COLUMNS = ['No.', 'Name', 'Position', 'Type', 'RPE', 'TR Time', 'Total Dist.', 'Dist. per min', 'HSR+Sprint', 'MSR', 'HSR', 'Sprint',
                           'Accel Cnt.', 'Decel Cnt.', 'Max Speed', 'GPS PL', 'Load']

    def __init__(self, PMANAGE):
        super().__init__(PMANAGE)
        self.output_excel_path = os.path.join(
            self.base_path, output_path)
        self.styler = ExcelStyle()
        self.chart = ExcelChart()

    def writeMatchExcel(self, matchCamp=None, matchDate='ALL', matchInfo=None):
        match_set = self.match_set
        position_row = 13
        player_row = 23

        if matchDate == 'ALL':
            file_path = os.path.join(
                self.output_excel_path, result_match)
            if os.path.isfile(file_path):
                os.remove(file_path)

            for (match_camp, match_date, match_info) in match_set.keys():
                match_excel, _, _ = self.matchTeamData(
                    match_camp, match_date, match_info)
                wtype, position, day = match_excel['type'], match_excel['position'][
                    self.MATCHPOSITION_COLUMNS], match_excel['day'][self.MATCHPLAYER_COLUMNS]
                day_type = match_date+'_'+match_info

                if os.path.isfile(file_path):
                    book = openpyxl.load_workbook(file_path)
                    writer = pd.ExcelWriter(
                        file_path, engine='openpyxl')
                    writer.book = book
                    position.to_excel(writer, sheet_name=day_type, index=False,
                                      float_format='%.1f', startrow=position_row, startcol=4)
                    day.to_excel(writer, sheet_name=day_type,
                                 index=False, float_format='%.1f', startrow=player_row)
                    writer.save()
                    writer.close()
                else:
                    writer = pd.ExcelWriter(
                        file_path, engine='openpyxl')
                    position.to_excel(writer, sheet_name=day_type, index=False,
                                      float_format='%.1f', startrow=position_row, startcol=4)
                    day.to_excel(writer, sheet_name=day_type,
                                 index=False, float_format='%.1f', startrow=player_row)
                    writer.save()
                    writer.close()
        else:
            match_excel, _, _ = self.matchTeamData(
                matchCamp, matchDate, matchInfo)
            file_path = os.path.join(
                self.output_excel_path, config.RESULT_CONFIG['match'], 'result_of_{}({}).xlsx'.format(matchDate, matchInfo))
            wtype, position, day = match_excel['type'], match_excel['position'][
                self.MATCHPOSITION_COLUMNS], match_excel['day'][self.MATCHPLAYER_COLUMNS]
            day_type = str(matchDate)+'_'+'M'
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            position.to_excel(writer, sheet_name=day_type, index=False,
                              float_format='%.1f', startrow=position_row, startcol=4)
            day.to_excel(writer, sheet_name=day_type,
                         index=False, float_format='%.1f', startrow=player_row)
            writer.save()
            writer.close()
        self.styler.styleDayExcel(file_path)
        self.chart.positionBar(file_path)
        return None

    def copyResultFile(self, copy_path):
        shutil.copy(os.path.join(
            self.base_path, output_path, result_player), copy_path)
        shutil.copy(os.path.join(
            self.base_path, output_path, result_day), copy_path)


class ExcelStyle():
    def stylePlayerExcel(self, file_path):
        player_excel_file = openpyxl.load_workbook(file_path)
        player_sheet_list = player_excel_file.sheetnames

        for sn in player_sheet_list:
            player_sheet = player_excel_file[sn]
            end_row = player_sheet.max_row
            end_col = player_sheet.max_column

            col_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I',
                        'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                        'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
            col_width = [17, 12, 8, 15, 5, 13, 13, 14, 13, 7,
                         7, 8, 8, 8, 16, 8, 7, 7, 9, 6, 9, 6, 9, 6,
                         8, 7, 7, 9, 6, 9, 6, 9, 6, 9, 6, 6, 6, 9, 6]

            for i in range(0, end_col):
                player_sheet.column_dimensions[col_name[i]
                                               ].width = col_width[i] + 2

                for j in range(2, end_row):
                    if j == 2:
                        top_cell = player_sheet.cell(j + 1, i + 1)
                        top_cell.fill = topFill
                        top_cell.alignment = centetAlignment
                        top_cell.border = normal_box
                    elif j == (end_row - 1):
                        bottom_cell = player_sheet.cell(end_row, i + 1)
                        bottom_cell.fill = bottomFill
                        bottom_cell.alignment = centetAlignment
                        bottom_cell.border = normal_box
                        bottom_cell.font = boldFont
                    else:
                        mid_cell = player_sheet.cell(j + 1, i + 1)
                        mid_cell.alignment = centetAlignment
                        mid_cell.border = normal_box

            for j in range(3, end_row - 1):
                front_cell = player_sheet.cell(j + 1, 1)
                front_cell.fill = frontFill

            player_sheet.merge_cells('A2:B2')
            player_sheet['A2'] = sn
            player_sheet['A2'].alignment = centetAlignment
            player_sheet['A2'].font = boldFont
            player_sheet['A2'].border = name_box
            player_sheet['B2'].border = name_box

        player_excel_file.save(file_path)

    def styleDayExcel(self, file_path):
        position_row = 13
        day_excel_file = openpyxl.load_workbook(file_path)
        day_sheet_list = day_excel_file.sheetnames

        for sn in day_sheet_list:
            p_num = 1

            day_sheet = day_excel_file[sn]

            end_row = day_sheet.max_row
            end_col = day_sheet.max_column

            col_name = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K',
                        'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
            col_width = [3, 14, 6, 8, 5, 13, 13, 14, 13, 7,
                         7, 8, 8, 8, 16, 8, 7, 7, 9, 6, 6, 6, 9, 6]

            ismatch = True if len(sn) > 8 else False

            if ismatch:
                position_list = ['Position', '1st half', '1st Half Avg.',
                                 '2nd half', '2nd half Avg.', 'Total', 'Total Avg.', 'Reserve Tot.', 'Reserve Avg.']
                for k in range(14, 23):
                    sc = 'A' + str(k)
                    ec = 'E' + str(k)
                    day_sheet.merge_cells(sc + ':' + ec)
                    day_sheet[sc] = position_list[k - 14]

                for i in range(position_row + 1, position_row + 9):
                    front_cell = day_sheet.cell(i + 1, 1)
                    if i >= position_row+7:
                        front_cell.fill = reserveFill
                    else:
                        front_cell.fill = frontFill

            else:
                position_list = ['Position']
                # position_list = ['Position', 'GK', 'CB', 'SB', 'MF', 'WF', 'FW', 'Team']

                k = 15
                pn = True
                while pn:
                    pc = 'E' + str(k)
                    pnm = day_sheet[pc].value
                    position_list.append(pnm)
                    if pnm == 'Team':
                        p_num += 1
                        pn = False
                    else:
                        k += 1
                        p_num += 1

                for k in range(14, 14+p_num):
                    sc = 'A' + str(k)
                    ec = 'E' + str(k)
                    day_sheet.merge_cells(sc + ':' + ec)
                    day_sheet[sc] = position_list[k - 14]

                for j in range(position_row + 1, position_row + p_num-1):
                    front_cell = day_sheet.cell(j + 1, 1)
                    front_cell.fill = frontFill

            # player_row = position_row + p_num + 2
            player_row = 23

            for i in range(0, end_col):
                day_sheet.column_dimensions[col_name[i]
                                            ].width = col_width[i] + 2

                for j in range(0, end_row):
                    day_sheet.row_dimensions[j + 1].height = 23
                    if j == position_row:
                        top_cell = day_sheet.cell(j + 1, i + 1)
                        top_cell.fill = topFill
                        top_cell.alignment = centetAlignment
                        top_cell.border = normal_box
                        top_cell.font = boldFont
                    elif j == position_row + p_num-1 and ismatch == False:
                        bottom_cell = day_sheet.cell(j + 1, i + 1)
                        bottom_cell.fill = bottomFill
                        bottom_cell.alignment = centetAlignment
                        bottom_cell.border = normal_box
                        bottom_cell.font = boldFont

                    elif j == player_row:
                        top_cell = day_sheet.cell(j + 1, i + 1)
                        top_cell.fill = topFill
                        top_cell.alignment = centetAlignment
                        top_cell.border = normal_box

                    elif j >= position_row and j < position_row + (p_num if ismatch == False else 9):
                        mid_cell = day_sheet.cell(j + 1, i + 1)
                        mid_cell.alignment = centetAlignment
                        mid_cell.border = normal_box
                    elif j > player_row:
                        mid_cell = day_sheet.cell(j + 1, i + 1)
                        mid_cell.alignment = centetAlignment
                        mid_cell.border = normal_box
            day_excel_file.save(file_path)


class ExcelChart():
    def positionBar(self, file_path):
        position_row = 13
        # player_row = 23
        chart_height = 9

        day_excel_file = openpyxl.load_workbook(file_path)
        day_sheet_list = day_excel_file.sheetnames

        for sn in day_sheet_list:
            ismatch = True if len(sn) > 8 else False

            day_sheet = day_excel_file[sn]

            if ismatch:
                pass
            else:
                k = 15
                p_num = 1
                pn = True
                while pn:
                    pc = 'A' + str(k)
                    if day_sheet[pc].value == 'Team':
                        p_num += 1
                        pn = False
                    else:
                        k += 1
                        p_num += 1

            chart1 = BarChart()

            chart1.type = "col"
            chart1.style = 10
            chart1.y_axis.title = 'Distance (m)'
            chart1.x_axis.title = 'Position'
            chart1.height = chart_height

            data1 = Reference(day_sheet, min_col=7, max_col=7, min_row=position_row +
                              2, max_row=position_row + (p_num if ismatch == False else 5))
            # labels = Reference(day_sheet, min_col=11, max_col=11, min_row=16, max_row=16)
            cats1 = Reference(day_sheet, min_col=1, min_row=position_row + 2,
                              max_row=position_row + (p_num if ismatch == False else 5))
            series = Series(data1, title='Total Dist.')
            chart1.append(series)
            chart1.set_categories(cats1)
            # chart1.layout = Layout(
            #     ManualLayout(
            #     x=0.25, y=0.25,
            #     h=0.5, w=0.5,
            #     xMode="edge",
            #     yMode="edge", til
            # ))
            chart1.legend.position = 't'
            # chart1.legend.layout = Layout(
            #     manualLayout=ManualLayout(
            #         yMode='edge',
            #         xMode='edge',
            #         x=0.2, y=0,
            #         h=0.1, w=0.5
            #     ))
            day_sheet.add_chart(chart1, "B2")

            chart2 = BarChart()

            chart2.type = "col"
            chart2.style = 10
            chart2.y_axis.title = 'Distance (m)'
            chart2.x_axis.title = 'Position'
            chart2.height = chart_height
            chart2.width = 22.5

            data2 = Reference(day_sheet, min_col=9, max_col=12, min_row=position_row +
                              1, max_row=position_row + (p_num if ismatch == False else 5))
            cats2 = Reference(day_sheet, min_col=1, min_row=position_row + 2,
                              max_row=position_row + (p_num if ismatch == False else 5))
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)
            chart2.shape = 10
            # chart2.layout = Layout(
            #     ManualLayout(
            #         x=0.25, y=0.25,
            #         h=0.5, w=0.5,
            #         xMode="edge",
            #         yMode="edge",
            #     ))
            chart2.legend.position = 't'
            # chart2.legend.layout = Layout(
            #     manualLayout=ManualLayout(
            #         yMode='edge',
            #         xMode='edge',
            #         x=0.2, y=0,
            #         h=0.1, w=0.5
            #     ))
            day_sheet.add_chart(chart2, "H2")

            chart3 = BarChart()

            chart3.type = "col"
            chart3.style = 10
            chart3.y_axis.title = 'Distance (m)'
            chart3.x_axis.title = 'Position'
            chart3.height = chart_height
            chart3.width = 17.5

            data3 = Reference(day_sheet, min_col=13, max_col=14, min_row=position_row +
                              1, max_row=position_row + (p_num if ismatch == False else 5))
            cats3 = Reference(day_sheet, min_col=1, min_row=position_row + 2,
                              max_row=position_row + (p_num if ismatch == False else 5))
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(cats3)
            chart3.shape = 10
            # chart3.layout = Layout(
            #     ManualLayout(
            #         x=0.25, y=0.25,
            #         h=0.5, w=0.5,
            #         xMode="edge",
            #         yMode="edge",
            #     ))
            chart3.legend.position = 't'
            # chart3.legend.layout = Layout(
            #     manualLayout=ManualLayout(
            #         yMode='edge',
            #         xMode='edge',
            #         x=0.2, y=0,
            #         h=0.1, w=0.5
            #     ))
            day_sheet.add_chart(chart3, "Q2")

        day_excel_file.save(file_path)
        day_excel_file.close()
