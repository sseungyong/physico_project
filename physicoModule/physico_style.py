import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font

boldFont = Font(bold=True)

centetAlignment = Alignment(horizontal='center', vertical='center')

normal_box = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
    diagonal=Side(border_style='thin', color='FF000000'),
    diagonal_direction=0,
    outline=Side(border_style='thin', color='FF000000'),
    vertical=Side(border_style='thin', color='FF000000'),
    horizontal=Side(border_style='thin', color='FF000000')
)

name_box = Border(
    bottom=Side(border_style='double', color='FF000000'),
)

topFill = PatternFill(patternType='solid', fgColor='fff200')
frontFill = PatternFill(patternType='solid', fgColor='dfe4ea')
bottomFill = PatternFill(patternType='solid', fgColor='ff6b6b')
reserveFill = PatternFill(patternType='solid', fgColor='b2bec3')
